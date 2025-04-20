import json
import csv
import openpyxl

def load_json(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        return json.load(file)

def load_csv(file_path):
    transactions = []
    with open(file_path, "r", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        for row in reader:
            transactions.append(row)
    return transactions

def load_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    transactions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        transactions.append(dict(zip(headers, row)))
    return transactions
